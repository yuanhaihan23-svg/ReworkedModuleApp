import pandas as pd
from datetime import datetime
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QPushButton, QTableWidget, QTableWidgetItem,
                             QHeaderView, QMessageBox,QComboBox,
                             QSplitter, QTextEdit, QFileDialog, QLineEdit,
                             QSizePolicy, QCompleter)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QBrush
from openpyxl.styles import PatternFill
from config import SCHEMA_FIELDS

# ----------------------------
# GUI CLASS
# ----------------------------
class BarcodeApp(QWidget):
    def __init__(self, db=None, whitelist=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.db = db
        self.current_barcode = None
        self.current_table = None
        self.session_barcodes = set()  # 本次运行的条码集合
        self.whitelist = whitelist

        # ----------------------
        # 窗口 & 总布局
        # ----------------------
        self.setWindowTitle("Rework Module Status Filling System")
        self.setGeometry(100, 50, 1400, 700)
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Barcode 输入行
        hbox = QHBoxLayout()
        layout.addLayout(hbox)

        # 标签
        hbox.addWidget(QLabel("Barcode:"))

        # 条码输入框
        self.barcode_edit = QLineEdit()
        self.barcode_edit.setFocus()
        self.barcode_edit.returnPressed.connect(self.scan_barcode)

        # 设置输入框随窗口横向拉伸，并指定最小宽度
        self.barcode_edit.setMinimumWidth(400)  # 最小宽度，可根据需求调大
        self.barcode_edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        hbox.addWidget(self.barcode_edit)

        # Scan 按钮
        scan_btn = QPushButton("Enter")
        scan_btn.clicked.connect(self.scan_barcode)
        hbox.addWidget(scan_btn)

        # ----------------------
        # 分割布局：左侧表单 / 右侧 overview + detail
        # ----------------------
        splitter = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(splitter)

        # 左侧表单
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        splitter.addWidget(left_widget)

        # 表格占据剩余空间
        self.form_table = QTableWidget()
        left_layout.addWidget(self.form_table, stretch=1)  # stretch=1 表示表格拉伸填满剩余空间

        # 提交按钮固定高度
        self.submit_btn = QPushButton("Submit")
        self.submit_btn.clicked.connect(self.submit_form)
        self.submit_btn.setEnabled(False)
        self.submit_btn.setFixedHeight(40)  # 可以根据需求调整高度
        left_layout.addWidget(self.submit_btn)  # stretch=0 默认固定高度

        # ----------------------
        # 右侧 Overview + Detail
        # ----------------------
        right_layout = QVBoxLayout()
        right_widget = QWidget()
        right_widget.setLayout(right_layout)
        splitter.addWidget(right_widget)

        # Overview 表格
        self.overview_table = QTableWidget()
        self.overview_table.setColumnCount(5)
        self.overview_table.setHorizontalHeaderLabels([
                        "Barcode",
            "Table1\nModule\nvisual status",
            "Table2\nCMM\ntest status",
            "Table3\nModule\nsampling status",
            "Overall\nstatus"
        ])

        self.overview_table.setWordWrap(True)  # 启用单元格文字换行
        self.overview_table.setAlternatingRowColors(True)  # 可选：隔行换色更易读
        self.overview_table.verticalHeader().setVisible(False)  # 隐藏行号

        # 按照扫码时间降序排列overview栏
        self.barcode_scan_time = {}  # {barcode: datetime对象}
        self.overview_table.cellClicked.connect(self.overview_cell_clicked)

        # 列宽控制：前 4 列可调节，最后一列自适应填满
        header = self.overview_table.horizontalHeader()
        for i in range(4):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)

        # 可选：允许用户拖拽调整列顺序
        header.setSectionsMovable(True)

        self.overview_table.cellClicked.connect(self.show_detail)
        right_layout.addWidget(self.overview_table)

        # ----------------------
        # 按钮区（刷新 / 导出）
        # ----------------------
        btn_box = QHBoxLayout()
        right_layout.addLayout(btn_box)

        refresh_btn = QPushButton("Refresh Overview")
        refresh_btn.clicked.connect(self.load_overview)
        btn_box.addWidget(refresh_btn)

        export_btn = QPushButton("Export Excel")
        export_btn.clicked.connect(self.export_excel_colored)
        btn_box.addWidget(export_btn)

        # ----------------------
        # Detail 显示区域
        # ----------------------
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        right_layout.addWidget(self.detail_text)

        # ----------------------
        # 初始化界面
        # ----------------------
        self.setup_ui()
        self.overview_table.setRowCount(0)
        self._apply_table_adjustments()

    # ----------------------
    # 自动调整行高
    # ----------------------
    def adjust_table_rows(self, table: QTableWidget):
        """根据内容自动调整行高"""
        table.resizeRowsToContents()

    # ----------------------
    # 左侧表单表格调整
    # ----------------------
    def adjust_form_table(self):
        """调整左侧表单表格列宽（标签列宽，值列窄）"""
        table = self.form_table
        if table.columnCount() == 0:
            return

        table_width = table.viewport().width()
        label_col_width = int(table_width * 0.75)
        value_col_width = table_width - label_col_width

        table.setColumnWidth(0, label_col_width)
        table.setColumnWidth(1, value_col_width)
        self.adjust_table_rows(table)

    # ----------------------
    # 右侧 overview 表格调整
    # ----------------------
    def adjust_overview_table(self):
        """调整右侧 Overview 表格列宽（条码宽，状态列略窄，总状态略宽）"""
        table = self.overview_table
        if table.columnCount() == 0:
            return

        total_width = table.viewport().width()
        proportions = [3, 1.2, 1.2, 1.2, 1]  # 可微调
        total_ratio = sum(proportions)

        for i, ratio in enumerate(proportions):
            col_width = int(total_width * ratio / total_ratio)
            table.setColumnWidth(i, col_width)

        self.adjust_table_rows(table)

    # ----------------------
    # 窗口大小变化时触发优化布局
    # ----------------------
    def resizeEvent(self, event):
        """根据窗口大小自动调整两侧表格的列宽与行高"""
        super().resizeEvent(event)

        # 使用 QTimer 防抖，避免频繁触发时卡顿或闪烁
        if hasattr(self, "_resize_timer"):
            self._resize_timer.stop()
        else:
            from PyQt6.QtCore import QTimer
            self._resize_timer = QTimer(self)
            self._resize_timer.setSingleShot(True)
            self._resize_timer.timeout.connect(self._apply_table_adjustments)

        self._resize_timer.start(100)  # 100ms 后调整，防止频繁调用

    def _apply_table_adjustments(self):
        """执行表格调整"""
        self.adjust_form_table()
        self.adjust_overview_table()

    def setup_ui(self):
        # 界面初始化代码
        pass

    def add_overview(self, status):
        if not self.current_barcode:
            QMessageBox.warning(self, "Warning", "Please scan a barcode first!")
            return
        # 保存到数据库
        self.db.add_overview(self.current_barcode, status)
        # 添加到 Overview 表格
        row_position = self.overview_table.rowCount()
        self.overview_table.insertRow(row_position)
        self.overview_table.setItem(row_position, 0, QTableWidgetItem(self.current_barcode))
        self.overview_table.setItem(row_position, 1, QTableWidgetItem(status))
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.overview_table.setItem(row_position, 2, QTableWidgetItem(timestamp))
        self.current_barcode = None  # 重置当前条码

    # ---------- 核心逻辑 ----------
    def scan_barcode(self):
        code = self.barcode_edit.text().strip()
        if not code:
            return

        # 校验条码长度
        if len(code) != 24:
            QMessageBox.warning(self, "Invalid barcode","Barcode must be 24 digits!")
            self.barcode_edit.clear()
            return

        self.current_barcode = code
        self.session_barcodes.add(code)
        from datetime import datetime
        self.barcode_scan_time[code] = datetime.now()

        # 确保数据库中有 overview 记录（如果没有则新建）
        self.db.ensure_barcode(code)

        # 实时检查每个子表是否存在
        has_t1 = bool(self.db.fetch_table("table_1_data", code))
        has_t2 = bool(self.db.fetch_table("table_2_data", code))
        has_t3 = bool(self.db.fetch_table("table_3_data", code))

        # 根据当前缺少的表单，自动决定要填写哪个
        if not has_t1:
            self.current_table = "table_1_data"
        elif not has_t2:
            self.current_table = "table_2_data"
        elif not has_t3:
            self.current_table = "table_3_data"
        else:
            self.current_table = None

        # 如果该条码已全部完成，则提示
        if not self.current_table:
            QMessageBox.information(self, "Note", f"Barcode\n{code} \nall tables already finished")
            self.form_table.setRowCount(0)
            self.submit_btn.setEnabled(False)
            return

        # 载入表单和 Overview
        self.load_form()
        self.load_overview()

    def load_form(self):
        self.form_table.clear()

        if not self.current_table:
            self.submit_btn.setEnabled(False)
            return

        fields = SCHEMA_FIELDS[self.current_table]
        self.form_table.setRowCount(len(fields))
        self.form_table.setColumnCount(2)
        self.form_table.setHorizontalHeaderLabels(["Inspection content", "Status"])

        for row_idx, (name, widget_type) in enumerate(fields):

            # 左列 label
            label = QTableWidgetItem(self.get_label(self.current_table, name))
            label.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.form_table.setItem(row_idx, 0, label)

            # ---------- timestamp 固定不可改 ----------
            if name == "timestamp":
                item = QTableWidgetItem(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                self.form_table.setItem(row_idx, 1, item)
                continue

            # ---------- quality_personnel 使用 lineedit ----------
            if name == "quality_personnel":
                edit = QLineEdit()
                history = self.db.fetch_quality_personnel_history()
                if history:
                    completer = QCompleter(history)
                    completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
                    completer.setFilterMode(Qt.MatchFlag.MatchContains)
                    edit.setCompleter(completer)
                self.form_table.setCellWidget(row_idx, 1, edit)
                continue
            DEFAULT_OPTIONS = {
                "default": ["OK", "NG"]
            }

            # ---------- 其他字段统一使用 QComboBox ----------
            combo = QComboBox()
            combo.addItems(DEFAULT_OPTIONS["default"])
            self.form_table.setCellWidget(row_idx, 1, combo)
        self.submit_btn.setEnabled(True)
        self._apply_table_adjustments()

    def highlight_widget(self, widget, item, color="#ffcccc"):
        if widget:
            widget.setStyleSheet(f"background-color: {color};")
        elif item:
            item.setBackground(QColor(*[int(c) for c in color.strip("#")[:6].replace("", "").split()]))

    def get_widget_value(self, widget, item):
        if isinstance(widget, QLineEdit):
            return widget.text().strip()
        elif isinstance(widget, QComboBox):
            return widget.currentText().strip()
        elif item:
            return item.text().strip()
        return ""

    def submit_form(self):
        if not self.current_table or not self.current_barcode:
            QMessageBox.warning(self, "Erro", "No forms available for submission")
            return

        missing_fields = []
        data = {}

        for i, (name, widget_type) in enumerate(SCHEMA_FIELDS[self.current_table]):
            item = self.form_table.item(i, 1)
            widget = self.form_table.cellWidget(i, 1)

            # ---------- timestamp ----------
            if name == "timestamp":
                data[name] = item.text() if item else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                continue

            # ---------- quality_personnel ----------
            if name == "quality_personnel":
                emp_id = widget.text().strip() if isinstance(widget, QLineEdit) else ""
                if not emp_id:
                    missing_fields.append(self.get_label(self.current_table, name))
                    self.highlight_widget(widget, item, "#ffcccc")
                    continue

                if emp_id not in self.whitelist:
                    QMessageBox.warning(
                        self,
                        "Unauthorized",
                        "❌Your employee ID is not on the list and cannot access this system"
                    )
                    self.highlight_widget(widget, item, "#ffcccc")
                    return

                data[name] = emp_id
                self.highlight_widget(widget, item, None)
                continue

            # ---------- 其他字段（统一下拉框处理） ----------
            value = ""
            if isinstance(widget, QComboBox):
                value = widget.currentText().strip()
            elif isinstance(widget, QLineEdit):
                value = widget.text().strip()
            elif item:
                value = item.text().strip()

            data[name] = value
            if not value:
                missing_fields.append(self.get_label(self.current_table, name))
                self.highlight_widget(widget, item, "#ffcccc")
            else:
                self.highlight_widget(widget, item, None)

        if missing_fields:
            QMessageBox.warning(
                self,
                "Required fields not filled out",
                "❌The following items are missing. Please complete them before submitting.\n\n" + "\n".join(missing_fields)
            )

            return

        # ---------- 写入数据库 ----------
        self.db.insert_or_update_table(self.current_table, self.current_barcode, data)

        # ---------- 更新表完成状态 ----------
        s1 = 1 if self.is_table_completed("table_1_data", self.current_barcode) else 0
        s2 = 1 if self.is_table_completed("table_2_data", self.current_barcode) else 0
        s3 = 1 if self.is_table_completed("table_3_data", self.current_barcode) else 0
        self.db.update_status(self.current_barcode, s1=s1, s2=s2, s3=s3)

        QMessageBox.information(
            self, "Successful submission", f"{self.get_table_name(self.current_table)} Done"
        )

        # ---------- 刷新 Overview ----------
        self.load_overview()

        # ---------- 清空当前表单 ----------
        self.current_barcode = None
        self.current_table = None
        self.barcode_edit.clear()
        self.barcode_edit.setFocus()
        self.form_table.setRowCount(0)
        self.submit_btn.setEnabled(False)
        self.detail_text.clear()

    def load_overview(self):
        """根据数据库真实填写情况计算 Overview 状态（字段完整度判断 + 新颜色逻辑）"""
        self.db._commit()

        try:
            if not self.session_barcodes:
                self.overview_table.setRowCount(0)
                return

            self.overview_table.setRowCount(0)

            # 按扫码时间排序
            barcodes_sorted = sorted(
                self.session_barcodes,
                key=lambda b: self.barcode_scan_time.get(b, datetime.min),
                reverse=True
            )

            for barcode in barcodes_sorted:

                # --------- 预读取子表数据 ---------
                t1_data = self.db.fetch_table("table_1_data", barcode) or {}
                t2_data = self.db.fetch_table("table_2_data", barcode) or {}
                t3_data = self.db.fetch_table("table_3_data", barcode) or {}

                # --------- 完成状态判断 ---------
                t1_done = self.is_table_completed("table_1_data", barcode)
                t2_done = self.is_table_completed("table_2_data", barcode)
                t3_done = self.is_table_completed("table_3_data", barcode)

                # --------- NG 判断函数 ---------
                def is_ng(data):
                    return any(str(v).strip().upper() == "NG" for v in data.values())

                # --------- 写入 overview ---------
                row = self.overview_table.rowCount()
                self.overview_table.insertRow(row)

                # 条码列
                barcode_item = QTableWidgetItem(barcode)
                barcode_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                self.overview_table.setItem(row, 0, barcode_item)

                # --------- 子表状态列处理 ---------
                table_datas = [t1_data, t2_data, t3_data]
                table_done_list = [t1_done, t2_done, t3_done]

                for col_index, (done, data) in enumerate(zip(table_done_list, table_datas), start=1):

                    if not done:  # 未完成
                        text = "Not Done"
                        bg = QColor(180, 180, 180)  # 灰色
                    else:
                        # 完成 → 判断是否 NG
                        if is_ng(data):
                            text = "NG"
                            bg = QColor(231, 76, 60)  # 红色
                        else:
                            text = "OK"
                            bg = QColor(46, 204, 113)  # 绿色

                    item = QTableWidgetItem(text)
                    item.setBackground(QBrush(bg))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                    self.overview_table.setItem(row, col_index, item)

                # --------- 最终状态列 ---------
                all_done = t1_done and t2_done and t3_done

                if not all_done:
                    final_text = "Not Finish"
                    final_bg = QColor(180, 180, 180)  # 灰色
                else:
                    # 所有表都完成 → 判断是否 NG
                    all_ng = is_ng(t1_data) or is_ng(t2_data) or is_ng(t3_data)
                    if all_ng:
                        final_text = "NG"
                        final_bg = QColor(231, 76, 60)  # 红色
                    else:
                        final_text = "OK"
                        final_bg = QColor(46, 204, 113)  # 绿色

                final_item = QTableWidgetItem(final_text)
                final_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                final_item.setBackground(QBrush(final_bg))
                final_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                self.overview_table.setItem(row, 4, final_item)

            self._apply_table_adjustments()

        except Exception as e:
            QMessageBox.critical(self, "Refresh erro", f"Erro happened while loading Overview: \n{str(e)}")

    # ---------- Detail 显示带颜色 ----------
    def show_detail(self,row,col):
        if row<0: return
        barcode=self.overview_table.item(row,0).text()
        html=f"<h2>Barcode: {barcode}</h2>"
        for tbl in SCHEMA_FIELDS:
            html+=f"<h3>{self.get_table_name(tbl)}</h3>"
            data=self.db.fetch_table(tbl,barcode)
            if not data:
                html+="<p><i>Not done</i></p>"
            else:
                html+="<table border='0' cellspacing='3' cellpadding='3'>"
                for k,v in data.items():
                    label=self.get_label(tbl,k)
                    val=str(v).strip()
                    color="#FF4C4C" if val.upper()=="NG" else "#2ECC71" if val.upper()=="OK" else "#000"
                    html+=f"<tr><td>{label}</td><td><b><font color='{color}'>{val}</font></b></td></tr>"
                html+="</table>"
            html+="<hr>"
        self.detail_text.setHtml(html)

    # ----------------------------
    # EXPORT EXCEL
    # ----------------------------
    def export_excel_colored(self):
        """导出当前 Overview 表格中显示的条码及对应的所有分表详细数据"""
        try:
            row_count = self.overview_table.rowCount()
            if row_count == 0:
                QMessageBox.warning(self, "No data", "There is no exportable data in the current table!")
                return

            # 选择保存路径
            default_name = f"current_detail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Choose file saved path", default_name, "Excel (*.xlsx)"
            )
            if not file_path:
                return

            # 提取当前 Overview 表格中所有条码
            barcodes = [self.overview_table.item(i, 0).text() for i in range(row_count)]

            # 从数据库获取 overview 数据
            all_rows = self.db.fetch_overview()
            rows = [r for r in all_rows if r[0] in barcodes]
            merged_df = pd.DataFrame(rows, columns=["barcode", "Table1_Module visual status", "Table2_CMM test status", "Table3_Module samping status", "Overall status"])

            # 合并所有分表
            for tbl, fields in SCHEMA_FIELDS.items():
                col_names = [f[0] for f in fields]
                cur = self.db._execute(f"SELECT barcode,{','.join(col_names)} FROM {tbl};")
                rows = cur.fetchall()
                df = pd.DataFrame(rows, columns=["barcode"] + col_names)
                merged_df = pd.merge(merged_df, df, on="barcode", how="left")

            # 只保留当前表格显示的条码
            merged_df = merged_df[merged_df["barcode"].isin(barcodes)]

            # 写入 Excel
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                merged_df.to_excel(writer, sheet_name="当前表格数据", index=False)
                ws = writer.sheets["当前表格数据"]

                # 定义颜色
                red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
                gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                # 给单元格上色
                for row in range(2, len(merged_df) + 2):
                    for col in range(2, merged_df.shape[1] + 1):
                        cell = ws.cell(row=row, column=col)
                        val = str(cell.value).upper() if cell.value else ""
                        if val in ["NG", "Not done"]:
                            cell.fill = red_fill
                        elif val in ["OK", "Done"]:
                            cell.fill = green_fill
                        elif not val:
                            cell.fill = gray_fill

            QMessageBox.information(self,"Export successful",f"The current table and detailed data have been exported:\n{file_path}")

        except Exception as e:QMessageBox.critical(self, "Export erro",f"An error occurred while exporting to Excel:\n{str(e)}")

    # ---------- Label 辅助 ----------
    def get_table_name(self,tbl):
        return {
            "table_1_data":"Table1_Module visual status",
            "table_2_data":"Table2_CMM test status",
            "table_3_data":"Table3_Module samping status"
        }.get(tbl,tbl)

    def get_label(self, tbl, field):
        labels = {
            "table_1_data": {
                "timestamp": "时间-Time",
                "quality_personnel": "品质人员-Quality",
                "protective_cover_installation_status": "正负极保护盖已安装-Positive and negative protective covers installed",
                "foreign_matters_status": "模组上保护盖无破损、无脏污-No damage or foreign matters on the module top cover",
                "protective_cover_appearance_status": "正负极连接铜排无破损，无异物-No damage or foreign matters on the positive and negative copper bar",
                "FPC_appearance_status": "FPC/卡扣无破损无脏污-No damage or foreign matters on FPC or buckles ",
                "FPC_foreign_matters_status": "FPC插接口无破损无脏污-No damage or foreign matters on FPC plug",
                "end_plate_hole_appearance_status": "端板孔无破损-No damage on end plate hole",
                "insulation_sheet_appearance_status": "端板上下绝缘片无破损-No damage on insulation",
                "side_plate_foam_appearance_status": "侧板缓冲棉清洁清洁-Clean of side plate cushion foam",
                "side_plate_appearance_status": "侧板与吊耳无划伤，无破损及变形-No scratch, damage or deformation on side plate and lugs",
                "bottom_glue_clean_status": "模组底部残胶无残留-No residue glue on module bottom",
                "module_bottom_picture": "模组底部照片-Module bottom pictures",
                "tray_foam_status": "托盘及缓冲泡棉无异物-No foreign matters on tray and cushion foam",
                "module_code_status": "模组码无破损脏污-No damage or dirty on module code"
            },
            "table_2_data": {
                "timestamp": "时间-Time",
                "quality_personnel": "品质人员-Quality",
                "module_CMMtest_status": "模组全尺寸测量-Module CMM test"
            },
            "table_3_data": {
                "module_CMMtest_status": "模组全尺寸测量-Module CMM test",
                "timestamp": "时间-Time",
                "quality_personnel": "品质人员-Quality",
                "sampling_data_status":"采样数据状态-Sampling data status",
                "voltage_status_consistent": "模组电压区间/充电未充电一致-Module voltage range/distinction between charging and uncharging",
                "cell_batch_consistent": "电芯批次一致-Cell batch consistency",
                "module_grade_consistent": "模组档位一致-Module grade consistency",
                "module_unbound_status": "模组已解绑-Module already unbounded",
                "module_appearance_status":"模组外观状态-Module appearance status"
            }
        }
        return labels.get(tbl, {}).get(field, field)

    def edit_barcode_from_table(self, row, col):
        barcode_item = self.overview_table.item(row, 0)
        if not barcode_item:
            return

        barcode = barcode_item.text()

        # 判断属于哪个表（table_1 → table_2 → table_3）
        for tbl in SCHEMA_FIELDS:
            if self.db.fetch_table(tbl, barcode):
                self.open_form(tbl, barcode)
                return

        # 默认新建进入 table_1
        self.open_form("table_1_data", barcode)

    def overview_cell_clicked(self, row, col):
        table_map = {1: "table_1_data", 2: "table_2_data", 3: "table_3_data"}

        barcode = self.overview_table.item(row, 0).text()

        if col in table_map:
            self.open_form(table_map[col], barcode)
        else:
            self.show_detail(row, col)

    def load_form_for_edit(self, table, barcode):
        """
        根据条码和表名加载表单，支持修改并重新提交
        """
        data = self.db.fetch_table(table, barcode)
        if not data:
            QMessageBox.information(self, "Note", f"{self.get_table_name(table)} No data available for editing.")
            return

        self.current_table = table
        self.current_barcode = barcode

        self.form_table.setRowCount(len(SCHEMA_FIELDS[table]))

        for i, (name, widget_type) in enumerate(SCHEMA_FIELDS[table]):
            # Label
            label_item = QTableWidgetItem(self.get_label(table, name))
            label_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.form_table.setItem(i, 0, label_item)

            # 值
            value = str(data.get(name, ""))

            if widget_type == "lineedit":
                line_edit = QLineEdit()
                line_edit.setText(value)
                self.form_table.setCellWidget(i, 1, line_edit)
            elif widget_type == "combobox":
                combo = QComboBox()
                combo.addItems(self.get_combobox_options(name))
                if value:
                    idx = combo.findText(value)
                    if idx >= 0:
                        combo.setCurrentIndex(idx)
                self.form_table.setCellWidget(i, 1, combo)
            else:
                item = QTableWidgetItem(value)
                self.form_table.setItem(i, 1, item)

        self.submit_btn.setEnabled(True)

    def open_form(self, table, barcode):
        """统一入口：判断权限 → 创建 UI → 填充数据"""

        # 权限检查：是否已完成？
        if not self.check_edit_permission(table, barcode):
            QMessageBox.warning(self, "Not editable",
                                f"{self.get_table_name(table)} is not yet done and cannot be edited!")
            return

        # ---- 合法才允许编辑 ----
        self.current_table = table
        self.current_barcode = barcode

        data = self.db.fetch_table(table, barcode) or {}

        self.load_form()
        self.fill_form(data)

        self.submit_btn.setEnabled(True)
        self.barcode_edit.setText(barcode)

    def fill_form(self, data):
        fields = SCHEMA_FIELDS[self.current_table]

        for row_idx, (name, _) in enumerate(fields):
            value = str(data.get(name, ""))

            widget = self.form_table.cellWidget(row_idx, 1)
            item = self.form_table.item(row_idx, 1)

            if name == "timestamp":
                continue

            if widget:
                if isinstance(widget, QLineEdit):
                    widget.setText(value)
                elif isinstance(widget, QComboBox):
                    idx = widget.findText(value)
                    if idx >= 0:
                        widget.setCurrentIndex(idx)
            elif item:
                item.setText(value)

    def is_table_completed(self, table, barcode):
        """判断指定表是否已完成（所有字段非空）"""
        data = self.db.fetch_table(table, barcode)
        if not data:
            return False

        for field, _ in SCHEMA_FIELDS[table]:
            if data.get(field, "") == "":
                return False

        return True

    def check_edit_permission(self, table, barcode):
        """已完成表单可修改；未完成表单不可修改"""
        return self.is_table_completed(table, barcode)

    def get_editable_tables(self, barcode):
        """根据填写进度返回当前条码允许编辑的表清单"""

        t1 = is_table_1_done = self.is_table_completed("table_1_data", barcode)
        t2 = is_table_2_done = self.is_table_completed("table_2_data", barcode)
        t3 = is_table_3_done = self.is_table_completed("table_3_data", barcode)

        # 规则：
        # 表3完成 → 可编辑 表1 表2
        if t3:
            return ["table_1_data", "table_2_data"]

        # 表2完成 → 可编辑 表1
        if t2:
            return ["table_1_data"]

        # 表1未完成 → 所有表可编辑
        return ["table_1_data", "table_2_data", "table_3_data"]







